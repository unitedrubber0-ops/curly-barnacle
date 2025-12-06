# edi.py
import pandas as pd, glob, os
from datetime import datetime, timedelta
from bs4 import BeautifulSoup
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

def build_schedule_df(html_folder=None, html_contents=None):
    """
    If html_folder is given, will glob *.html inside it.
    Alternatively, you can directly pass html_contents as a list of strings.
    Returns a DataFrame with your columns + data.
    """
    # 1) determine week columns
    today = pd.Timestamp.now().normalize()
    start_monday = today - pd.Timedelta(days=today.weekday())
    weeks = pd.date_range(start=start_monday, periods=52, freq='W-MON')
    WEEK_COLS = [d.strftime('%Y-%m-%d') for d in weeks]
    OUTPUT_COLS = ['PLANT','PART','Route','Project'] + WEEK_COLS + ['Grand Total']

    # 2) load html
    html_list = []
    if html_contents:
        html_list = html_contents
    elif html_folder:
        for fn in glob.glob(os.path.join(html_folder, '*.html')):
            with open(fn, 'r', encoding='utf-8') as f:
                html_list.append(f.read())
    else:
        raise ValueError("Need either html_folder or html_contents")

    # 3) parse each file and build rows
    rows = []
    for html in html_list:
        soup = BeautifulSoup(html, 'html.parser')
        # --- your existing parsing code here; for each row append:
        # rows.append([plant, part, route, project] + week_qtys + [sum(week_qtys)])
        #
        # For demo: a dummy zero‑row
        rows.append(['Plant A','Part 123','R1','Proj X'] + [0]*52 + [0])

    return pd.DataFrame(rows, columns=OUTPUT_COLS)


def df_to_excel_bytes(df):
    """Turn a DataFrame into an in‑memory .xlsx file (BytesIO)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Schedule"

    # header
    bold = Font(bold=True)
    for i, col in enumerate(df.columns, start=1):
        cell = ws.cell(row=1, column=i, value=col)
        cell.font = bold
        cell.alignment = Alignment(horizontal='center')

    # data rows
    for r_idx, row in enumerate(df.itertuples(index=False), start=2):
        for c_idx, val in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=val)

    # auto‑width
    for i, col_cells in enumerate(ws.columns, start=1):
        maxlen = max(len(str(c.value)) for c in col_cells if c.value is not None)
        ws.column_dimensions[get_column_letter(i)].width = maxlen + 2

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio
