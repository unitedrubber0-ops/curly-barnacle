import io
import os
import re
import math
import glob
import tempfile
import logging
from datetime import datetime, timedelta
from collections import defaultdict

from flask import (
    Flask, request, send_file,
    render_template, flash,
    redirect, url_for
)
import pandas as pd
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.utils.datetime import from_excel

# ——— Logging ————————————————————————————————————————————————————————
logging.basicConfig(level=logging.DEBUG, format='%(levelname)s: %(message)s')

app = Flask(__name__)
app.secret_key = os.getenv("SECRET_KEY", "replace-with-secure-random-key")

# ——— Helper 1: Conditional formatting on a schedule workbook ——————————————————
def apply_conditional_formatting(
    input_path,
    output_path,
    sheet_name: str = "Schedule",
    part_col_header: str = "PART",
):
    wb = load_workbook(input_path)
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb[wb.sheetnames[0]]

    # ── 1. Locate Headers ──────────────────────────────────────────────────────
    header_row = next(
        (
            c.row
            for r in ws.iter_rows(min_row=1, max_row=20)
            for c in r
            if c.value and re.search(r"\bStock\b", str(c.value), re.IGNORECASE)
        ),
        None,
    )
    if header_row is None:
        raise ValueError("Header row with 'Stock' not found.")

    headers = [c.value for c in ws[header_row]]
    
    # Identify key columns
    try:
        stock_col = next(
            i for i, h in enumerate(headers, 1) 
            if h and re.search(r"\bStock\b", str(h), re.IGNORECASE)
        )
        part_col = next(
            i for i, h in enumerate(headers, 1) 
            if h and re.search(part_col_header, str(h), re.IGNORECASE)
        )
    except StopIteration:
        raise ValueError("Could not find 'Stock' or Part column headers.")
    
    # Identify IT columns (IT01, IT02...)
    it_col_indices = [
        i for i, h in enumerate(headers, 1) 
        if h and re.match(r"IT\d+", str(h), re.IGNORECASE)
    ]

    # ── 2. Parse Dates (Flexible Parser) ──────────────────────────────────────
    def parse_date(v):
        if v is None: return None
        try:
            if isinstance(v, datetime): return v.date()
        except: pass
        if isinstance(v, (int, float)):
            try: return from_excel(v).date()
            except: pass
        
        s = str(v).strip().strip('"\'')
        if not s: return None
        
        for fmt in ("%d-%m-%Y", "%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%y%m%d", "%Y%m%d"):
            try: return datetime.strptime(s, fmt).date()
            except: continue
        
        try:
            pd_dt = pd.to_datetime(s, dayfirst=True, errors="coerce")
            if not pd.isna(pd_dt): return pd_dt.date()
        except: pass
        return None

    it_cols = {} 
    for idx in it_col_indices:
        d = parse_date(headers[idx-1]) or parse_date(ws.cell(row=header_row, column=idx).value)
        if not d and header_row > 1:
            d = parse_date(ws.cell(row=header_row - 1, column=idx).value)
        it_cols[idx] = d or datetime.min.date()

    req_cols = {} 
    for idx, h in enumerate(headers, 1):
        if idx not in it_col_indices and idx != stock_col:
            d = parse_date(h)
            if d:
                req_cols[idx] = d

    sorted_reqs = sorted(req_cols.items(), key=lambda x: x[1])

    # ── 3. Define Styles ──────────────────────────────────────────────────────
    green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    yellow = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    red = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    
    # ── 4. Group Rows by Part ─────────────────────────────────────────────────
    parts = defaultdict(list)
    for r in range(header_row + 1, ws.max_row + 1):
        part_val = ws.cell(row=r, column=part_col).value
        if part_val is not None:
            parts[part_val].append(r)

    # ── 5. Processing Phase (Calculate & Store in Memory) ─────────────────────
    
    # We store calculations here first because we need to insert columns later,
    # which changes cell indices.
    row_meta = {}      # {row_idx: {'G': date, 'Y': date, 'R': date, 'wh': int, 'it': int}}
    cell_colors = {}   # {row_idx: {old_col_idx: color_obj}}

    def get_val(r, c):
        v = ws.cell(r, c).value
        if v is None: return 0.0
        try:
            return float(v)
        except:
            try:
                return float(str(v).replace(',', '').strip())
            except:
                return 0.0

    for part, rows in parts.items():
        wh_stock = max(get_val(r, stock_col) for r in rows)
        
        it_supply = []
        for col, date in it_cols.items():
            qty = max(get_val(r, col) for r in rows) 
            if qty > 0:
                it_supply.append({'date': date, 'qty': qty})
        it_supply.sort(key=lambda x: x['date'])

        current_wh = wh_stock
        current_it_pool = 0.0 
        it_ptr = 0 
        
        # Determine Color Logic
        last_row_color = {r: green for r in rows}
        
        # Track First Dates (Part Level or Row Level - User asked for Row Level dates)
        # We will track first occurrences for each row individually
        row_dates = {r: {'Y': None, 'R': None} for r in rows}

        current_part_status = green

        for col_idx, req_date in sorted_reqs:
            # Update Pool
            while it_ptr < len(it_supply) and it_supply[it_ptr]['date'] <= req_date:
                current_it_pool += it_supply[it_ptr]['qty']
                it_ptr += 1

            # Demand Check
            row_values = {r: get_val(r, col_idx) for r in rows}
            weekly_demand = sum(row_values.values())

            if weekly_demand > 0:
                if current_wh >= weekly_demand:
                    current_wh -= weekly_demand
                    current_part_status = green
                else:
                    gap = weekly_demand - current_wh
                    current_wh = 0
                    if current_it_pool >= gap:
                        current_it_pool -= gap
                        current_part_status = yellow 
                    else:
                        current_it_pool = 0
                        current_part_status = red
            
            date_str = req_date.strftime("%Y-%m-%d")

            # Apply Logic to Rows
            for r in rows:
                val = row_values[r]
                final_color = None
                
                if val > 0:
                    final_color = current_part_status
                    last_row_color[r] = current_part_status
                else:
                    final_color = last_row_color[r]
                
                # Store color for later application
                if r not in cell_colors: cell_colors[r] = {}
                cell_colors[r][col_idx] = final_color

                # Capture First Date
                if final_color == yellow and row_dates[r]['Y'] is None:
                    row_dates[r]['Y'] = date_str
                elif final_color == red and row_dates[r]['R'] is None:
                    row_dates[r]['R'] = date_str

        # Calculate Metrics (Months Coverage)
        wh_calc = max(get_val(r, stock_col) for r in rows)
        it_total_calc = sum(max(get_val(r, col) for r in rows) for col in it_cols.keys())
        total_supply = wh_calc + it_total_calc
        demands = [sum(get_val(r, col) for r in rows) for col, _ in sorted_reqs]
        
        def calc_coverage(stock, demand_list):
            months = 0
            for d in demand_list:
                if d == 0: continue 
                if stock >= d:
                    stock -= d
                    months += 1
                else:
                    break
            return months

        wh_months = calc_coverage(wh_calc, demands)
        all_months = calc_coverage(total_supply, demands)

        # Save Metadata
        for r in rows:
            row_meta[r] = {
                'Y': row_dates[r]['Y'],
                'R': row_dates[r]['R'],
                'cov_wh': wh_months,
                'cov_it': all_months
            }

    # ── 6. Modification Phase (Insert Columns) ────────────────────────────────
    # We want to insert BEFORE the first demand column
    if req_cols:
        insert_idx = min(req_cols.keys())
    else:
        insert_idx = ws.max_column + 1

    # Insert 4 columns: First Y, First R, Cov WH, Cov IT
    ws.insert_cols(insert_idx, 4)

    # Write Headers
    titles = ["First Yellow Date", "First Red Date", "Cov (WH)", "Cov (WH+IT)"]
    for i, t in enumerate(titles):
        c = ws.cell(row=header_row, column=insert_idx + i)
        c.value = t
        c.font = Font(bold=True)

    # ── 7. Write Data & Apply Formatting ──────────────────────────────────────
    for r in range(header_row + 1, ws.max_row + 1):
        # A. Write New Columns
        if r in row_meta:
            m = row_meta[r]
            ws.cell(r, insert_idx).value = m['Y'] or "-"
            ws.cell(r, insert_idx+1).value = m['R'] or "-"
            ws.cell(r, insert_idx+2).value = m['cov_wh']
            ws.cell(r, insert_idx+3).value = m['cov_it']

        # B. Color Demand Cells (Shifted by 4)
        if r in cell_colors:
            for old_col, color in cell_colors[r].items():
                # If the column was at or after insertion point, shift it
                target_col = old_col + 4 if old_col >= insert_idx else old_col
                ws.cell(r, target_col).fill = color

    # ── 8. Autofit Columns ────────────────────────────────────────────────────
    for i in range(1, ws.max_column + 1):
        col_letter = get_column_letter(i)
        max_len = 0
        # Check first 100 rows to speed up
        for r in range(1, min(ws.max_row, 100) + 1):
            val = ws.cell(r, i).value
            if val:
                try:
                    length = len(str(val))
                    if length > max_len: max_len = length
                except: pass
        
        # Add padding
        adjusted_width = (max_len + 2) * 1.1
        ws.column_dimensions[col_letter].width = max(adjusted_width, 8) # Min width 8

    wb.save(output_path)
    logging.info(f"Saved updated coverage report to {output_path}")


# ——— Helper 2: aggregate a single "Merged"‑sheet EDI snapshot —————————————
def aggregate_edi_from_path(path):
    TARGET = 'Merged'
    now = pd.Timestamp.now().normalize()
    # next three calendar months
    starts = [now.replace(day=1) + pd.DateOffset(months=i) for i in (1,2,3)]
    ends   = [s + pd.DateOffset(months=1) - pd.Timedelta(days=1) for s in starts]
    labels = [s.strftime('%Y-%m') for s in starts]

    xls = pd.ExcelFile(path)
    sheets = [s for s in xls.sheet_names if TARGET.lower() in s.lower()]
    if not sheets:
        raise ValueError(f"No '{TARGET}' sheet in {os.path.basename(path)}")
    sheet = sheets[0]

    raw = pd.read_excel(xls, sheet_name=sheet, header=None)
    hdr = None
    for idx,row in raw.iterrows():
        if any(isinstance(c,str) and c.strip().lower()=='plant' for c in row):
            hdr = idx
            break
    if hdr is None:
        raise ValueError(f"Header row not found in {os.path.basename(path)}")

    df = pd.read_excel(xls, sheet_name=sheet, header=hdr)
    df.columns = [str(c).strip() for c in df.columns]
    df.fillna(0, inplace=True)
    if 'Part No' in df: df.rename(columns={'Part No':'PART'}, inplace=True)
    if 'Plant'   in df: df.rename(columns={'Plant':'PLANT'},   inplace=True)

    weeks = [c for c in df.columns
             if isinstance(c,str) and pd.notna(pd.to_datetime(c, errors='coerce'))]

    ids = ['PLANT','PART'] + [c for c in ('Route','Project') if c in df]
    melt = df.melt(id_vars=ids, value_vars=weeks,
                   var_name='Week', value_name='Qty')
    melt['Week'] = pd.to_datetime(melt['Week'], errors='coerce')
    melt.dropna(subset=['Week'], inplace=True)

    def bucket(d):
        for lab,s,e in zip(labels, starts, ends):
            if s <= d <= e:
                return lab
        return None

    melt['Month'] = melt['Week'].apply(bucket)
    filt = melt.dropna(subset=['Month'])
    summary = (filt
               .pivot_table(index=ids, columns='Month',
                            values='Qty', aggfunc='sum',
                            fill_value=0)
               .reset_index())
    # ensure all three months present
    for lab in labels:
        if lab not in summary:
            summary[lab] = 0
    return summary, labels


# ——— Task 1: HTML → Schedule.xlsx —————————————————————————————————————
@app.route('/', methods=['GET'])
def index():
    return render_template('index.html')

@app.route('/convert', methods=['GET', 'POST'])
def convert():
    if request.method == 'GET':
        return render_template('convert.html')

    files = request.files.getlist('html_files')
    if not files or not any(f.filename.lower().endswith('.html') for f in files):
        flash('Upload at least one .html file.')
        return redirect(url_for('convert'))

    # Build weekly columns
    today = pd.Timestamp.now().normalize()
    start = today - pd.Timedelta(days=today.weekday())
    weeks = pd.date_range(start, periods=52, freq='W-MON')
    WEEKCOLS = [d.strftime('%Y-%m-%d') for d in weeks]
    OUTCOLS = ['PLANT', 'PART', 'Route', 'Project'] + WEEKCOLS + ['Grand Total']

    rows = []
    for f in files:
        try:
            html = f.stream.read().decode('utf-8')
            soup = BeautifulSoup(html, 'html.parser')

            # Extract plant
            plant = ''
            b_from = soup.find('b', string=lambda t: t and t.strip() == 'From:')
            if b_from:
                sib = b_from.next_sibling
                if sib and isinstance(sib, str) and sib.strip():
                    plant = sib.strip()
                else:
                    plant = b_from.parent.get_text().replace('From:', '').strip()

            # Extract parts and forecasts
            for part_td in soup.find_all(
                    'td', string=lambda t: t and "Buyer's Part Number:" in t):
                # Part number
                next_td = part_td.find_next_sibling('td')
                part = next_td.get_text(strip=True) if next_td else ''

                data_week = {wk: 0 for wk in WEEKCOLS}
                # Locate forecast table
                table = None
                for tbl in part_td.find_all_next('table'):
                    if tbl.find('div', attrs={'date': True}):
                        table = tbl
                        break

                if table:
                    for tr in table.find_all('tr')[1:]:
                        cells = tr.find_all('td')
                        if len(cells) < 3:
                            continue
                        qty_text = cells[0].get_text(strip=True)
                        date_div = cells[2].find('div', attrs={'date': True})
                        if not qty_text or date_div is None:
                            continue
                        try:
                            qty = float(qty_text)
                        except ValueError:
                            continue
                        date_str = date_div['date']
                        dt = None
                        for fmt in ('%y%m%d', '%Y%m%d'):
                            try:
                                dt = datetime.strptime(date_str, fmt).date()
                                break
                            except ValueError:
                                continue
                        if not dt:
                            continue
                        monday = dt - timedelta(days=dt.weekday())
                        wk = monday.strftime('%Y-%m-%d')
                        if wk in data_week:
                            data_week[wk] += qty

                grand_total = sum(data_week.values())
                rows.append({
                    'PLANT': plant,
                    'PART': part,
                    'Route': '',
                    'Project': '',
                    **data_week,
                    'Grand Total': grand_total
                })
        except Exception as e:
            logging.warning(f"Error parsing HTML {f.filename}: {e}")

    if not rows:
        flash('No data found in HTML files.')
        return redirect(url_for('convert'))

    # Build Excel workbook
    df = pd.DataFrame(rows, columns=OUTCOLS)
    buf = io.BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = 'Merged'

    bold = Font(bold=True)
    for idx, col in enumerate(OUTCOLS, start=1):
        cell = ws.cell(row=1, column=idx, value=col)
        cell.font = bold
        cell.alignment = Alignment(horizontal='center')

    for r_idx, tup in enumerate(df.itertuples(index=False), start=2):
        for c_idx, val in enumerate(tup, start=1):
            ws.cell(row=r_idx, column=c_idx, value=val)

    ws.freeze_panes = 'A2'
    for i, col_cells in enumerate(ws.columns, start=1):
        max_len = max(len(str(c.value)) if c.value else 0 for c in col_cells)
        ws.column_dimensions[get_column_letter(i)].width = max_len + 2

        # Save workbook to buffer
    wb.save(buf)
    # Return workbook directly
    buf.seek(0)
    return send_file(
        buf,
        download_name='EDI_Schedule.xlsx',
        as_attachment=True,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

# ——— Task 2: Coverage Report —————————————————————————————————————
@app.route("/coverage", methods=("GET","POST"))
def coverage():
    if request.method=="GET":
        return render_template("coverage.html")

    f = request.files.get("schedule_file")
    if not f or not f.filename.lower().endswith((".xlsx",".xlsm",".xltx",".xltm")):
        flash("Valid Excel required"); return redirect(url_for("coverage"))

    tmp_in = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    tmp_in.write(f.read()); tmp_in.flush()
    tmp_out = tmp_in.name.replace(".xlsx","_coverage.xlsx")

    apply_conditional_formatting(tmp_in.name, tmp_out,
                                 sheet_name=request.form.get("sheet_name","Schedule"),
                                 part_col_header=request.form.get("part_header","PART"))
    return send_file(tmp_out,
                     download_name="Coverage_Report.xlsx",
                     as_attachment=True)


# ——— Task 3: EDI Fluctuations —————————————————————————————————————
@app.route("/fluctuations", methods=["GET", "POST"])
def fluctuations():
    if request.method == "GET":
        return render_template("fluctuations.html")

    uploaded = request.files.getlist("edi_files")
    paths = []
    for f in uploaded:
        if f and f.filename.lower().endswith(".xlsx"):
            tf = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
            tf.write(f.read()); tf.flush()
            paths.append(tf.name)

    if len(paths) < 3:
        flash("Please upload at least three .xlsx files.")
        return redirect(url_for("fluctuations"))

    # sort by modification time → oldest, mid, latest
    paths.sort(key=lambda p: os.path.getmtime(p))
    old_path, mid_path, latest_path = paths[-3], paths[-2], paths[-1]

    try:
        # aggregate each snapshot
        old_df, months    = aggregate_edi_from_path(old_path)
        mid_df, _         = aggregate_edi_from_path(mid_path)
        latest_df, _      = aggregate_edi_from_path(latest_path)

        # suffix month columns
        for df_snap, tag in ((old_df, "_old"),
                             (mid_df, "_mid"),
                             (latest_df, "_latest")):
            df_snap.rename(
                columns={m: f"{m}{tag}" for m in months},
                inplace=True
            )

        # build merge keys: only extras present in all three
        keys = ["PLANT", "PART"]
        for extra in ("Route", "Project"):
            if (extra in old_df.columns
                and extra in mid_df.columns
                and extra in latest_df.columns):
                keys.append(extra)

        # outer‑merge all three snapshots
        merged = (
            old_df
            .merge(mid_df,    on=keys, how="outer")
            .merge(latest_df, on=keys, how="outer")
            .fillna(0)
        )

        # totals & fluctuation
        for prefix in ("old", "mid", "latest"):
            cols = [f"{m}_{prefix}" for m in months]
            merged[f"Total_{prefix}"] = merged[cols].sum(axis=1)

        def pct_change(row):
            old_val    = row["Total_old"]
            latest_val = row["Total_latest"]
            if old_val == 0:
            # both zero → 0%, otherwise infinite growth → cap at 100%
               return 0 if latest_val == 0 else 100
        # normal % change
            return (latest_val - old_val) / old_val * 100

        merged["Fluctuation_%"] = merged.apply(pct_change, axis=1)

        # write to Excel in-memory
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            merged.to_excel(writer, sheet_name="Fluctuation", index=False)
        output.seek(0)

        return send_file(
            output,
            download_name="EDI_Fluctuation_Report.xlsx",
            as_attachment=True,
            mimetype=(
              "application/vnd.openxmlformats-officedocument"
              ".spreadsheetml.sheet"
            )
        )

    finally:
        for p in paths:
            try: os.unlink(p)
            except: pass


# ——— Helper: scan for critical (unmet) parts and build a .xlsx —————————————
def generate_unmet_requirements_excel(
    input_xlsx: str,
    sheet_name: str = 'Schedule',
    window_weeks: int = 8,
    date_formats=(' %d-%m-%Y', '%Y-%m-%d', '%y%m%d', '%Y%m%d')
) -> io.BytesIO:
    wb = load_workbook(input_xlsx, data_only=True)
    ws = wb[sheet_name]

    # 1) Find header row by "Stock"
    header_row = next(
        (cell.row
         for row in ws.iter_rows(min_row=1, max_row=20)
         for cell in row
         if cell.value and re.search(r'\bStock\b', str(cell.value), re.IGNORECASE)),
        None
    )
    if header_row is None:
        raise ValueError("Header row containing 'Stock' not found.")

    # 2) Read header values
    headers = [cell.value for cell in ws[header_row]]

    # 3) Flexible date parser
    def parse_date(v):
        # a) already datetime
        if isinstance(v, datetime):
            return v
        # b) Excel serial number
        if isinstance(v, (int, float)):
            try:
                return from_excel(v)
            except:
                pass
        # c) string formats
        for fmt in date_formats:
            try:
                return datetime.strptime(str(v).strip(), fmt)
            except:
                continue
        return None

    # 4) Collect date columns
    parsed = [(idx, parse_date(hdr))
              for idx, hdr in enumerate(headers, start=1)]
    # only keep those we parsed successfully
    parsed_dates = [(i,d) for i,d in parsed if d]
    if not parsed_dates:
        raise ValueError("No date‑formatted headers found.")

    # window
    start_date = min(d for _,d in parsed_dates)
    end_date   = start_date + timedelta(weeks=window_weeks)
    req_date_cols = [(i,d) for i,d in parsed_dates
                     if start_date <= d <= end_date]

    # 5) Find PART and Plant columns
    part_col  = next(i for i,h in enumerate(headers,1)
                     if h and re.search(r'\bPART\b', str(h), re.IGNORECASE))
    plant_col = next(i for i,h in enumerate(headers,1)
                     if h and re.search(r'\bPlant\b', str(h), re.IGNORECASE))

    # 6) Red‑fill detector
    def is_red(cell):
        rgb = getattr(getattr(cell.fill, 'fgColor', None), 'rgb', '')
        return 'FFC7CE' in (rgb or '').upper()

    # 7) Scan rows for unmet (robust numeric parsing)
    data = []
    def parse_float_safe(v):
        try:
            return float(v)
        except Exception:
            try:
                return float(str(v).replace(',', '').strip())
            except Exception:
                return 0.0

    for row in ws.iter_rows(min_row=header_row+1, max_row=ws.max_row):
        part  = row[part_col-1].value
        plant = row[plant_col-1].value
        for col_idx, req_date in req_date_cols:
            cell = ws.cell(row=row[0].row, column=col_idx)
            if is_red(cell) and cell.value and parse_float_safe(cell.value) > 0:
                data.append({
                    'Part Number': part,
                    'Plant': plant,
                    'Requirement Date': req_date.date().isoformat(),
                    'Unmet Qty': parse_float_safe(cell.value)
                })

    # 8) Build and return Excel
    df = pd.DataFrame(data, columns=[
        'Part Number', 'Plant', 'Requirement Date', 'Unmet Qty'
    ])

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Critical_Parts')
    output.seek(0)
    return output


# ——— Route: Critical Parts —————————————————————————————————————
@app.route("/critical_parts", methods=["GET", "POST"])
def critical_parts():
    if request.method == "GET":
        return render_template("critical_parts.html")

    uploaded = request.files.get("schedule_file")
    logging.debug(f"Uploaded file: {uploaded and uploaded.filename}")
    if not uploaded or not uploaded.filename.lower().endswith((".xlsx", ".xlsm")):
        flash("Please upload a valid Excel schedule file.")
        return redirect(url_for("critical_parts"))

    # write temp
    tf = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    tf.write(uploaded.read()); tf.flush(); tf.close()
    try:
        excel_io = generate_unmet_requirements_excel(
            input_xlsx=tf.name,
            sheet_name=request.form.get("sheet_name", "Schedule"),
            window_weeks=int(request.form.get("window_weeks", 8))
        )
        logging.debug(f"Generated Excel; size={excel_io.getbuffer().nbytes} bytes")
    except Exception as e:
        logging.exception("Failed to generate critical parts Excel")
        flash(f"Error: {e}")
        return redirect(url_for("critical_parts"))
    finally:
        os.unlink(tf.name)

    excel_io.seek(0)
    return send_file(
        excel_io,
        as_attachment=True,
        download_name="Critical_Parts_Report.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


@app.context_processor
def inject_current_year():
    return {'current_year': datetime.now().year}

if __name__=="__main__":
    app.run(debug=True)
